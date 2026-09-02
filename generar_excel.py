import os
import pandas as pd
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from extractor_mlb import obtener_partidos_dia
from extractor_odds import obtener_momios_f5

def calcular_probabilidad_f5(row):
    fip_vis, fip_loc = row['FIP_Vis'], row['FIP_Loc']
    whip_vis, whip_loc = row['WHIP_Vis'], row['WHIP_Loc']
    
    prob_local = 54.0
    prob_local += ((fip_vis - fip_loc) * 14.0)
    prob_local += ((whip_vis - whip_loc) * 8.0)
    
    v_vis, v_loc = row['V_Vis_L10'], row['V_Loc_L10']
    
    if v_loc >= 7: prob_local += 3.0
    elif v_loc <= 3: prob_local -= 3.0
        
    if v_vis >= 7: prob_local -= 3.0
    elif v_vis <= 3: prob_local += 3.0
    
    prob_local = max(30.0, min(80.0, prob_local))
    prob_visita = 100.0 - prob_local
    
    if prob_local >= prob_visita:
        return row['Local'], round(prob_local, 1)
    else:
        return row['Visitante'], round(prob_visita, 1)

def crear_reporte_excel(fecha=None):
    if fecha is None:
        fecha = datetime.today().strftime('%Y-%m-%d')
        
    df = obtener_partidos_dia(fecha)
    if df is None or df.empty:
        print("No hay datos para exportar.")
        return

    print("Obteniendo momios en tiempo real...")
    dict_odds = obtener_momios_f5()

    resultados = []
    for _, row in df.iterrows():
        equipo_f5, prob_est = calcular_probabilidad_f5(row)
        
        datos_odds = dict_odds.get(equipo_f5, {'momio_amer': -110, 'momio_dec': 1.91})
        momio_amer = datos_odds['momio_amer']
        momio_dec = datos_odds['momio_dec']
        
        prob_imp = round((1 / momio_dec) * 100, 1)
        ev_val = round(prob_est - prob_imp, 1)

        resultados.append({
            'Hora': row['Hora'],
            'Visitante': row['Visitante'],
            'L10_Vis': row['Racha_Vis'],
            'P_Visita': row['P_Visita'],
            'FIP_Vis': row['FIP_Vis'],
            'WHIP_Vis': row['WHIP_Vis'],
            'Local': row['Local'],
            'L10_Loc': row['Racha_Loc'],
            'P_Local': row['P_Local'],
            'FIP_Loc': row['FIP_Loc'],
            'WHIP_Loc': row['WHIP_Loc'],
            'Seleccion_F5': equipo_f5,
            'Prob_Modelo_%': prob_est,
            'Momio': momio_amer if momio_amer else "-110",
            'Momio_Dec': momio_dec,
            'Prob_Casa_%': prob_imp,
            'EV_%': ev_val
        })
    
    df_res = pd.DataFrame(resultados).sort_values(by='EV_%', ascending=False)
    nombre_archivo = "MLB_F5_Reporte.xlsx"
    
    if os.path.exists(nombre_archivo):
        wb = load_workbook(nombre_archivo)
    else:
        wb = Workbook()
        wb.remove(wb.active)

    if fecha in wb.sheetnames:
        del wb[fecha]
        
    ws = wb.create_sheet(title=fecha)

    headers = ["Hora (UTC)", "Visitante", "L10 Vis", "Pitcher Visita", "FIP Vis", "WHIP Vis", 
               "Local", "L10 Loc", "Pitcher Local", "FIP Loc", "WHIP Loc", "Recomendación F5", 
               "Prob Modelo (%)", "Momio Amer", "Prob Casa (%)", "+EV (%)"]
    ws.append(headers)

    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    
    green_fill = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    
    thin_border = Border(
        left=Side(style='thin', color='D3D3D3'), right=Side(style='thin', color='D3D3D3'),
        top=Side(style='thin', color='D3D3D3'), bottom=Side(style='thin', color='D3D3D3')
    )

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for _, row in df_res.iterrows():
        row_data = [
            row['Hora'], row['Visitante'], row['L10_Vis'], row['P_Visita'], row['FIP_Vis'], row['WHIP_Vis'],
            row['Local'], row['L10_Loc'], row['P_Local'], row['FIP_Loc'], row['WHIP_Loc'],
            row['Seleccion_F5'], row['Prob_Modelo_%'], row['Momio'], row['Prob_Casa_%'], row['EV_%']
        ]
        ws.append(row_data)
        current_row = ws.max_row
        ev_val = row['EV_%']
        
        for col_num in range(1, len(row_data) + 1):
            cell = ws.cell(row=current_row, column=col_num)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center" if col_num in [1, 3, 5, 6, 8, 10, 11, 13, 14, 15, 16] else "left")
            
            if ev_val >= 8.0:
                cell.fill = green_fill
            elif ev_val >= 3.0:
                cell.fill = yellow_fill

    # --- RECUADRO RESUMEN: PARLAY DEL DÍA (TOP 3 +EV) ---
    top_3 = df_res.head(3)
    if len(top_3) == 3:
        ws.append([]) # Fila vacía
        row_start = ws.max_row + 1
        
        ws.cell(row=row_start, column=1, value="PARLAY SUGERIDO DEL DÍA (TOP 3 +EV)").font = Font(bold=True, size=11, color="FFFFFF")
        ws.cell(row=row_start, column=1).fill = PatternFill(start_color="203764", end_color="203764", fill_type="solid")
        ws.merge_cells(start_row=row_start, start_column=1, end_row=row_start, end_column=4)
        
        cuota_parlay = 1.0
        prob_parlay = 1.0
        
        for idx, (_, item) in enumerate(top_3.iterrows(), start=1):
            r = row_start + idx
            ws.cell(row=r, column=1, value=f"Selección {idx}:")
            ws.cell(row=r, column=2, value=f"{item['Seleccion_F5']} (F5)")
            ws.cell(row=r, column=3, value=f"Momio: {item['Momio']}")
            ws.cell(row=r, column=4, value=f"+EV: {item['EV_%']}%")
            
            cuota_parlay *= item['Momio_Dec']
            prob_parlay *= (item['Prob_Modelo_%'] / 100.0)
        
        r_final = row_start + 4
        ws.cell(row=r_final, column=1, value="Cuota Decimal Total:").font = Font(bold=True)
        ws.cell(row=r_final, column=2, value=round(cuota_parlay, 2)).font = Font(bold=True)
        ws.cell(row=r_final, column=3, value="Prob. Acierto Conjunta:").font = Font(bold=True)
        ws.cell(row=r_final, column=4, value=f"{round(prob_parlay * 100, 1)}%").font = Font(bold=True)

    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = col[0].column_letter
        ws.column_dimensions[col_letter].width = max(max_len + 3, 11)

    wb.save(nombre_archivo)
    print(f"\n ¡Reporte final guardado en '{fecha}' con resumen de apuestas!")

if __name__ == "__main__":
    crear_reporte_excel()